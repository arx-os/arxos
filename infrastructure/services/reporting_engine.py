"""
Advanced Reporting Engine Service.

Provides comprehensive report generation capabilities including PDF, Excel,
CSV, and HTML formats with dynamic templates, charts, and data visualizations.
"""

import asyncio
from typing import Dict, Any, List, Optional, Union, BinaryIO
from datetime import datetime, timezone
from pathlib import Path
import json
import uuid
import tempfile
import shutil

# Import for PDF generation
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# Import for Excel generation
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.chart import LineChart, BarChart, PieChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Import for chart generation
try:
    import matplotlib.pyplot as plt
    import matplotlib.dates as mdates
    from matplotlib.backends.backend_agg import FigureCanvasAgg
    import seaborn as sns
    import pandas as pd
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False

from domain.entities.analytics_entity import ReportTemplate, ReportFormat, ReportStatus
from infrastructure.logging.structured_logging import get_logger
from infrastructure.storage.file_storage import FileStorageService


logger = get_logger(__name__)


class ReportingEngineService:
    """Advanced reporting engine for generating various types of reports."""
    
    def __init__(self, storage_service: FileStorageService = None):
        self.storage_service = storage_service or FileStorageService()
        self.temp_dir = Path(tempfile.mkdtemp())
        
        # Report generators by format
        self.generators = {
            ReportFormat.PDF: self._generate_pdf_report,
            ReportFormat.EXCEL: self._generate_excel_report,
            ReportFormat.CSV: self._generate_csv_report,
            ReportFormat.JSON: self._generate_json_report,
            ReportFormat.HTML: self._generate_html_report
        }
        
        # Chart generators
        self.chart_generators = {
            "line": self._generate_line_chart,
            "bar": self._generate_bar_chart,
            "pie": self._generate_pie_chart,
            "scatter": self._generate_scatter_chart,
            "histogram": self._generate_histogram
        }
        
        # Setup matplotlib style if available
        if MATPLOTLIB_AVAILABLE:
            plt.style.use('seaborn-v0_8')
            sns.set_palette("husl")
    
    async def generate_report(self, template: ReportTemplate, data: Dict[str, Any], 
                            format: ReportFormat, options: Dict[str, Any] = None) -> Dict[str, Any]:
        """Generate report based on template and data."""
        report_id = str(uuid.uuid4())
        start_time = datetime.now(timezone.utc)
        
        try:
            logger.info(f"Starting report generation: {report_id}")
            
            # Validate inputs
            if format not in self.generators:
                raise ValueError(f"Unsupported report format: {format}")
            
            # Prepare report context
            context = self._prepare_report_context(template, data, options or {})
            
            # Generate report
            generator = self.generators[format]
            file_path, metadata = await generator(template, context, report_id)
            
            # Store report file
            stored_file_info = await self._store_report_file(file_path, report_id, format)
            
            # Calculate generation metrics
            end_time = datetime.now(timezone.utc)
            generation_time_ms = int((end_time - start_time).total_seconds() * 1000)
            
            # Prepare response
            result = {
                "report_id": report_id,
                "format": format.value,
                "status": ReportStatus.COMPLETED.value,
                "file_path": str(file_path),
                "download_url": stored_file_info.get("download_url"),
                "file_size": stored_file_info.get("file_size", 0),
                "page_count": metadata.get("page_count", 1),
                "generation_time_ms": generation_time_ms,
                "generated_at": end_time.isoformat(),
                "metadata": metadata
            }
            
            logger.info(f"Report generation completed: {report_id} in {generation_time_ms}ms")
            return result
            
        except Exception as e:
            logger.error(f"Report generation failed for {report_id}: {e}")
            return {
                "report_id": report_id,
                "status": ReportStatus.FAILED.value,
                "error": str(e),
                "generated_at": datetime.now(timezone.utc).isoformat()
            }
    
    def _prepare_report_context(self, template: ReportTemplate, data: Dict[str, Any], 
                              options: Dict[str, Any]) -> Dict[str, Any]:
        """Prepare report generation context."""
        context = {
            "template": template,
            "data": data,
            "options": options,
            "generated_at": datetime.now(timezone.utc),
            "charts": {},
            "tables": {},
            "metadata": {
                "template_name": template.name,
                "template_version": template.version,
                "data_sources": template.data_sources
            }
        }
        
        # Generate charts if data available
        if "chart_data" in data and MATPLOTLIB_AVAILABLE:
            context["charts"] = self._generate_charts(data["chart_data"], options)
        
        return context
    
    async def _generate_pdf_report(self, template: ReportTemplate, context: Dict[str, Any], 
                                 report_id: str) -> tuple[Path, Dict[str, Any]]:
        """Generate PDF report."""
        if not REPORTLAB_AVAILABLE:
            raise ImportError("ReportLab is required for PDF generation")
        
        file_path = self.temp_dir / f"{report_id}.pdf"
        
        # Create PDF document
        doc = SimpleDocTemplate(str(file_path), pagesize=letter)
        story = []
        styles = getSampleStyleSheet()
        
        # Add title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=24,
            spaceAfter=30,
            textColor=colors.darkblue
        )
        story.append(Paragraph(template.name, title_style))
        story.append(Spacer(1, 20))
        
        # Add generation info
        gen_info = f"Generated on: {context['generated_at'].strftime('%Y-%m-%d %H:%M:%S UTC')}"
        story.append(Paragraph(gen_info, styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Process template sections
        for section in template.sections:
            await self._add_pdf_section(story, section, context, styles)
        
        # Build PDF
        doc.build(story)
        
        # Get file size and page count
        file_size = file_path.stat().st_size
        
        metadata = {
            "page_count": len(story),  # Approximate
            "file_size": file_size,
            "sections_count": len(template.sections)
        }
        
        return file_path, metadata
    
    async def _add_pdf_section(self, story: List, section: Dict[str, Any], 
                             context: Dict[str, Any], styles) -> None:
        """Add section to PDF story."""
        section_type = section["type"]
        title = section["title"]
        config = section["config"]
        
        # Add section title
        story.append(Paragraph(title, styles['Heading2']))
        story.append(Spacer(1, 10))
        
        if section_type == "executive_summary":
            await self._add_executive_summary_pdf(story, context, styles)
        elif section_type == "building_metrics":
            await self._add_building_metrics_pdf(story, context, styles)
        elif section_type == "data_table":
            await self._add_data_table_pdf(story, context, config, styles)
        elif section_type == "chart":
            await self._add_chart_pdf(story, context, config)
        
        story.append(Spacer(1, 20))
    
    async def _add_executive_summary_pdf(self, story: List, context: Dict[str, Any], styles) -> None:
        """Add executive summary section to PDF."""
        data = context["data"].get("executive_summary", {})
        
        # Key metrics summary
        if "key_metrics" in data:
            metrics = data["key_metrics"]
            story.append(Paragraph("Key Metrics:", styles['Heading3']))
            
            metrics_data = [
                ["Metric", "Value"],
                ["Total Buildings", str(metrics.get("total_buildings", 0))],
                ["Operational Buildings", str(metrics.get("operational_buildings", 0))],
                ["Total Floors", str(metrics.get("total_floors", 0))],
                ["Total Rooms", str(metrics.get("total_rooms", 0))],
                ["Total Devices", str(metrics.get("total_devices", 0))]
            ]
            
            metrics_table = Table(metrics_data)
            metrics_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(metrics_table)
            story.append(Spacer(1, 10))
        
        # Highlights
        if "highlights" in data:
            story.append(Paragraph("Highlights:", styles['Heading3']))
            for highlight in data["highlights"]:
                story.append(Paragraph(f"• {highlight}", styles['Normal']))
            story.append(Spacer(1, 10))
        
        # Concerns
        if "concerns" in data:
            story.append(Paragraph("Areas of Concern:", styles['Heading3']))
            for concern in data["concerns"]:
                story.append(Paragraph(f"• {concern}", styles['Normal']))
    
    async def _add_building_metrics_pdf(self, story: List, context: Dict[str, Any], styles) -> None:
        """Add building metrics section to PDF."""
        data = context["data"].get("building_metrics", {})
        
        # Create metrics table
        metrics_data = [["Metric", "Value", "Unit"]]
        
        for key, value in data.items():
            if isinstance(value, (int, float)):
                unit = ""
                if "percentage" in key or "rate" in key:
                    unit = "%"
                elif "cost" in key:
                    unit = "$"
                elif "temperature" in key:
                    unit = "°C"
                
                metrics_data.append([key.replace("_", " ").title(), str(value), unit])
        
        if len(metrics_data) > 1:
            table = Table(metrics_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.navy),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightblue),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
            ]))
            story.append(table)
    
    async def _add_data_table_pdf(self, story: List, context: Dict[str, Any], 
                                config: Dict[str, Any], styles) -> None:
        """Add data table to PDF."""
        table_data = config.get("data", [])
        if not table_data:
            return
        
        # Create table
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(table)
    
    async def _add_chart_pdf(self, story: List, context: Dict[str, Any], config: Dict[str, Any]) -> None:
        """Add chart to PDF."""
        chart_type = config.get("type", "line")
        chart_data = config.get("data", {})
        
        if not chart_data or not MATPLOTLIB_AVAILABLE:
            return
        
        # Generate chart
        chart_path = await self._generate_chart(chart_type, chart_data, config)
        if chart_path:
            img = Image(str(chart_path), width=6*inch, height=4*inch)
            story.append(img)
    
    async def _generate_excel_report(self, template: ReportTemplate, context: Dict[str, Any], 
                                   report_id: str) -> tuple[Path, Dict[str, Any]]:
        """Generate Excel report."""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel generation")
        
        file_path = self.temp_dir / f"{report_id}.xlsx"
        workbook = openpyxl.Workbook()
        
        # Remove default sheet
        workbook.remove(workbook.active)
        
        # Create sheets based on template sections
        for section in template.sections:
            await self._add_excel_sheet(workbook, section, context)
        
        # Save workbook
        workbook.save(str(file_path))
        
        metadata = {
            "sheet_count": len(workbook.worksheets),
            "file_size": file_path.stat().st_size
        }
        
        return file_path, metadata
    
    async def _add_excel_sheet(self, workbook, section: Dict[str, Any], context: Dict[str, Any]) -> None:
        """Add sheet to Excel workbook."""
        section_type = section["type"]
        title = section["title"]
        
        worksheet = workbook.create_sheet(title=title[:31])  # Excel sheet name limit
        
        if section_type == "building_metrics":
            await self._add_building_metrics_excel(worksheet, context)
        elif section_type == "data_table":
            await self._add_data_table_excel(worksheet, context, section["config"])
        elif section_type == "executive_summary":
            await self._add_executive_summary_excel(worksheet, context)
    
    async def _add_building_metrics_excel(self, worksheet, context: Dict[str, Any]) -> None:
        """Add building metrics to Excel sheet."""
        data = context["data"].get("building_metrics", {})
        
        # Headers
        worksheet['A1'] = "Metric"
        worksheet['B1'] = "Value"
        worksheet['C1'] = "Unit"
        
        # Apply header formatting
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for col in ['A1', 'B1', 'C1']:
            worksheet[col].font = header_font
            worksheet[col].fill = header_fill
        
        # Data rows
        row = 2
        for key, value in data.items():
            if isinstance(value, (int, float)):
                worksheet[f'A{row}'] = key.replace("_", " ").title()
                worksheet[f'B{row}'] = value
                
                # Determine unit
                unit = ""
                if "percentage" in key or "rate" in key:
                    unit = "%"
                elif "cost" in key:
                    unit = "$"
                elif "temperature" in key:
                    unit = "°C"
                
                worksheet[f'C{row}'] = unit
                row += 1
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    async def _add_data_table_excel(self, worksheet, context: Dict[str, Any], config: Dict[str, Any]) -> None:
        """Add data table to Excel sheet."""
        table_data = config.get("data", [])
        if not table_data:
            return
        
        # Write data
        for row_idx, row_data in enumerate(table_data, 1):
            for col_idx, cell_value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
                
                # Format header row
                if row_idx == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    async def _add_executive_summary_excel(self, worksheet, context: Dict[str, Any]) -> None:
        """Add executive summary to Excel sheet."""
        data = context["data"].get("executive_summary", {})
        row = 1
        
        # Report period
        if "report_period" in data:
            period = data["report_period"]
            worksheet[f'A{row}'] = "Report Period"
            worksheet[f'B{row}'] = f"{period.get('start', '')} to {period.get('end', '')}"
            row += 2
        
        # Key metrics
        if "key_metrics" in data:
            worksheet[f'A{row}'] = "Key Metrics"
            worksheet[f'A{row}'].font = Font(bold=True)
            row += 1
            
            metrics = data["key_metrics"]
            for key, value in metrics.items():
                worksheet[f'A{row}'] = key.replace("_", " ").title()
                worksheet[f'B{row}'] = value
                row += 1
            
            row += 1
        
        # Highlights
        if "highlights" in data:
            worksheet[f'A{row}'] = "Highlights"
            worksheet[f'A{row}'].font = Font(bold=True)
            row += 1
            
            for highlight in data["highlights"]:
                worksheet[f'A{row}'] = f"• {highlight}"
                row += 1
            
            row += 1
        
        # Concerns
        if "concerns" in data:
            worksheet[f'A{row}'] = "Areas of Concern"
            worksheet[f'A{row}'].font = Font(bold=True)
            row += 1
            
            for concern in data["concerns"]:
                worksheet[f'A{row}'] = f"• {concern}"
                row += 1
    
    async def _generate_csv_report(self, template: ReportTemplate, context: Dict[str, Any], 
                                 report_id: str) -> tuple[Path, Dict[str, Any]]:
        """Generate CSV report."""
        import csv
        
        file_path = self.temp_dir / f"{report_id}.csv"
        
        with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            
            # Write header with report info
            writer.writerow([f"Report: {template.name}"])
            writer.writerow([f"Generated: {context['generated_at'].strftime('%Y-%m-%d %H:%M:%S UTC')}"])
            writer.writerow([])  # Empty row
            
            # Process each section
            for section in template.sections:
                await self._add_csv_section(writer, section, context)
                writer.writerow([])  # Empty row between sections
        
        metadata = {
            "file_size": file_path.stat().st_size,
            "encoding": "utf-8"
        }
        
        return file_path, metadata
    
    async def _add_csv_section(self, writer, section: Dict[str, Any], context: Dict[str, Any]) -> None:
        """Add section to CSV."""
        section_type = section["type"]
        title = section["title"]
        
        writer.writerow([f"=== {title} ==="])
        
        if section_type == "building_metrics":
            data = context["data"].get("building_metrics", {})
            writer.writerow(["Metric", "Value"])
            for key, value in data.items():
                if isinstance(value, (int, float)):
                    writer.writerow([key.replace("_", " ").title(), value])
        
        elif section_type == "data_table":
            table_data = section["config"].get("data", [])
            for row in table_data:
                writer.writerow(row)
    
    async def _generate_json_report(self, template: ReportTemplate, context: Dict[str, Any], 
                                  report_id: str) -> tuple[Path, Dict[str, Any]]:
        """Generate JSON report."""
        file_path = self.temp_dir / f"{report_id}.json"
        
        # Prepare JSON structure
        report_json = {
            "report_id": report_id,
            "template": {
                "name": template.name,
                "type": template.type,
                "version": template.version
            },
            "generated_at": context["generated_at"].isoformat(),
            "data": context["data"],
            "sections": []
        }
        
        # Process sections
        for section in template.sections:
            section_data = {
                "type": section["type"],
                "title": section["title"],
                "config": section["config"],
                "data": context["data"].get(section["type"], {})
            }
            report_json["sections"].append(section_data)
        
        # Write JSON file
        with open(file_path, 'w', encoding='utf-8') as jsonfile:
            json.dump(report_json, jsonfile, indent=2, ensure_ascii=False, default=str)
        
        metadata = {
            "file_size": file_path.stat().st_size,
            "encoding": "utf-8",
            "sections_count": len(report_json["sections"])
        }
        
        return file_path, metadata
    
    async def _generate_html_report(self, template: ReportTemplate, context: Dict[str, Any], 
                                  report_id: str) -> tuple[Path, Dict[str, Any]]:
        """Generate HTML report."""
        file_path = self.temp_dir / f"{report_id}.html"
        
        # Generate HTML content
        html_content = self._build_html_content(template, context)
        
        # Write HTML file
        with open(file_path, 'w', encoding='utf-8') as htmlfile:
            htmlfile.write(html_content)
        
        metadata = {
            "file_size": file_path.stat().st_size,
            "encoding": "utf-8"
        }
        
        return file_path, metadata
    
    def _build_html_content(self, template: ReportTemplate, context: Dict[str, Any]) -> str:
        """Build HTML content for report."""
        html = f"""
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>{template.name}</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 40px; line-height: 1.6; }}
                h1 {{ color: #333; border-bottom: 2px solid #007bff; padding-bottom: 10px; }}
                h2 {{ color: #555; margin-top: 30px; }}
                table {{ border-collapse: collapse; width: 100%; margin: 20px 0; }}
                th, td {{ border: 1px solid #ddd; padding: 12px; text-align: left; }}
                th {{ background-color: #f5f5f5; font-weight: bold; }}
                .metric-value {{ font-weight: bold; color: #007bff; }}
                .section {{ margin-bottom: 30px; }}
                .generated-info {{ font-style: italic; color: #666; margin-bottom: 30px; }}
            </style>
        </head>
        <body>
            <h1>{template.name}</h1>
            <div class="generated-info">
                Generated on: {context['generated_at'].strftime('%Y-%m-%d %H:%M:%S UTC')}
            </div>
        """
        
        # Process sections
        for section in template.sections:
            html += f'<div class="section">\n<h2>{section["title"]}</h2>\n'
            html += self._build_html_section(section, context)
            html += "</div>\n"
        
        html += """
        </body>
        </html>
        """
        
        return html
    
    def _build_html_section(self, section: Dict[str, Any], context: Dict[str, Any]) -> str:
        """Build HTML for a section."""
        section_type = section["type"]
        
        if section_type == "building_metrics":
            return self._build_metrics_table_html(context["data"].get("building_metrics", {}))
        elif section_type == "executive_summary":
            return self._build_executive_summary_html(context["data"].get("executive_summary", {}))
        elif section_type == "data_table":
            return self._build_data_table_html(section["config"].get("data", []))
        
        return "<p>Section data not available</p>"
    
    def _build_metrics_table_html(self, data: Dict[str, Any]) -> str:
        """Build metrics table HTML."""
        html = "<table>\n<thead>\n<tr><th>Metric</th><th>Value</th></tr>\n</thead>\n<tbody>\n"
        
        for key, value in data.items():
            if isinstance(value, (int, float)):
                metric_name = key.replace("_", " ").title()
                html += f'<tr><td>{metric_name}</td><td class="metric-value">{value}</td></tr>\n'
        
        html += "</tbody>\n</table>\n"
        return html
    
    def _build_executive_summary_html(self, data: Dict[str, Any]) -> str:
        """Build executive summary HTML."""
        html = ""
        
        # Highlights
        if "highlights" in data:
            html += "<h3>Highlights</h3>\n<ul>\n"
            for highlight in data["highlights"]:
                html += f"<li>{highlight}</li>\n"
            html += "</ul>\n"
        
        # Concerns
        if "concerns" in data:
            html += "<h3>Areas of Concern</h3>\n<ul>\n"
            for concern in data["concerns"]:
                html += f"<li>{concern}</li>\n"
            html += "</ul>\n"
        
        return html
    
    def _build_data_table_html(self, table_data: List[List[str]]) -> str:
        """Build data table HTML."""
        if not table_data:
            return "<p>No data available</p>"
        
        html = "<table>\n"
        
        # Header row
        if table_data:
            html += "<thead>\n<tr>"
            for cell in table_data[0]:
                html += f"<th>{cell}</th>"
            html += "</tr>\n</thead>\n"
        
        # Data rows
        html += "<tbody>\n"
        for row in table_data[1:]:
            html += "<tr>"
            for cell in row:
                html += f"<td>{cell}</td>"
            html += "</tr>\n"
        html += "</tbody>\n"
        
        html += "</table>\n"
        return html
    
    def _generate_charts(self, chart_data: Dict[str, Any], options: Dict[str, Any]) -> Dict[str, str]:
        """Generate charts and return paths."""
        charts = {}
        
        if not MATPLOTLIB_AVAILABLE:
            return charts
        
        for chart_name, chart_config in chart_data.items():
            chart_type = chart_config.get("type", "line")
            data = chart_config.get("data", {})
            
            if data and chart_type in self.chart_generators:
                try:
                    chart_path = self.chart_generators[chart_type](data, chart_config)
                    if chart_path:
                        charts[chart_name] = str(chart_path)
                except Exception as e:
                    logger.warning(f"Failed to generate chart {chart_name}: {e}")
        
        return charts
    
    async def _generate_chart(self, chart_type: str, data: Dict[str, Any], config: Dict[str, Any]) -> Optional[Path]:
        """Generate a single chart."""
        if not MATPLOTLIB_AVAILABLE or chart_type not in self.chart_generators:
            return None
        
        try:
            return self.chart_generators[chart_type](data, config)
        except Exception as e:
            logger.warning(f"Failed to generate {chart_type} chart: {e}")
            return None
    
    def _generate_line_chart(self, data: Dict[str, Any], config: Dict[str, Any]) -> Optional[Path]:
        """Generate line chart."""
        if not data.get("x") or not data.get("y"):
            return None
        
        fig, ax = plt.subplots(figsize=(10, 6))
        ax.plot(data["x"], data["y"], marker='o', linewidth=2, markersize=6)
        
        ax.set_xlabel(config.get("x_label", "X"))
        ax.set_ylabel(config.get("y_label", "Y"))
        ax.set_title(config.get("title", "Line Chart"))
        ax.grid(True, alpha=0.3)
        
        chart_path = self.temp_dir / f"chart_{uuid.uuid4().hex}.png"
        plt.savefig(chart_path, dpi=300, bbox_inches='tight')
        plt.close()
        
        return chart_path
    
    def _generate_bar_chart(self, data: Dict[str, Any], config: Dict[str, Any]) -> Optional[Path]:
        """Generate bar chart."""
        if not data.get("categories") or not data.get("values"):
            return None
        
        fig, ax = plt.subplots(figsize=(10, 6))
        bars = ax.bar(data["categories"], data["values"])
        
        # Add value labels on bars
        for bar in bars:
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2., height,
                   f'{height:.1f}', ha='center', va='bottom')
        
        ax.set_xlabel(config.get("x_label", "Categories"))
        ax.set_ylabel(config.get("y_label", "Values"))
        ax.set_title(config.get("title", "Bar Chart"))
        
        # Rotate x-axis labels if needed
        plt.xticks(rotation=45, ha='right')
        
        chart_path = self.temp_dir / f"chart_{uuid.uuid4().hex}.png"
        plt.savefig(chart_path, dpi=300, bbox_inches='tight')
        plt.close()
        
        return chart_path
    
    def _generate_pie_chart(self, data: Dict[str, Any], config: Dict[str, Any]) -> Optional[Path]:
        """Generate pie chart."""
        if not data.get("labels") or not data.get("values"):
            return None
        
        fig, ax = plt.subplots(figsize=(8, 8))
        ax.pie(data["values"], labels=data["labels"], autopct='%1.1f%%', startangle=90)
        ax.set_title(config.get("title", "Pie Chart"))
        
        chart_path = self.temp_dir / f"chart_{uuid.uuid4().hex}.png"
        plt.savefig(chart_path, dpi=300, bbox_inches='tight')
        plt.close()
        
        return chart_path
    
    def _generate_scatter_chart(self, data: Dict[str, Any], config: Dict[str, Any]) -> Optional[Path]:
        """Generate scatter chart."""
        if not data.get("x") or not data.get("y"):
            return None
        
        fig, ax = plt.subplots(figsize=(10, 6))
        ax.scatter(data["x"], data["y"], alpha=0.7, s=50)
        
        ax.set_xlabel(config.get("x_label", "X"))
        ax.set_ylabel(config.get("y_label", "Y"))
        ax.set_title(config.get("title", "Scatter Plot"))
        ax.grid(True, alpha=0.3)
        
        chart_path = self.temp_dir / f"chart_{uuid.uuid4().hex}.png"
        plt.savefig(chart_path, dpi=300, bbox_inches='tight')
        plt.close()
        
        return chart_path
    
    def _generate_histogram(self, data: Dict[str, Any], config: Dict[str, Any]) -> Optional[Path]:
        """Generate histogram."""
        if not data.get("values"):
            return None
        
        fig, ax = plt.subplots(figsize=(10, 6))
        ax.hist(data["values"], bins=config.get("bins", 20), alpha=0.7, edgecolor='black')
        
        ax.set_xlabel(config.get("x_label", "Values"))
        ax.set_ylabel(config.get("y_label", "Frequency"))
        ax.set_title(config.get("title", "Histogram"))
        ax.grid(True, alpha=0.3)
        
        chart_path = self.temp_dir / f"chart_{uuid.uuid4().hex}.png"
        plt.savefig(chart_path, dpi=300, bbox_inches='tight')
        plt.close()
        
        return chart_path
    
    async def _store_report_file(self, file_path: Path, report_id: str, format: ReportFormat) -> Dict[str, Any]:
        """Store report file and return storage info."""
        if self.storage_service:
            try:
                storage_key = f"reports/{report_id}.{format.value}"
                stored_info = await self.storage_service.upload_file(str(file_path), storage_key)
                return stored_info
            except Exception as e:
                logger.warning(f"Failed to store report file: {e}")
        
        return {
            "file_size": file_path.stat().st_size,
            "local_path": str(file_path)
        }
    
    def cleanup_temp_files(self) -> None:
        """Clean up temporary files."""
        try:
            shutil.rmtree(self.temp_dir)
            self.temp_dir = Path(tempfile.mkdtemp())
        except Exception as e:
            logger.warning(f"Failed to cleanup temp files: {e}")
    
    def __del__(self):
        """Cleanup on destruction."""
        try:
            self.cleanup_temp_files()
        except:
            pass